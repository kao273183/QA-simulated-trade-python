import re,json,os,argparse
import openpyxl
import requests as rs
from datetime import datetime,timedelta

# 1 // 未平倉
# 2 // 平倉
# 3 // 止盈平倉
# 4 // 止損平倉
# 5 // 爆倉/強制平倉

#domain = 'http://pre.top.one'
headers = {
        'Content-Type':'application/json',
    }
#參數
balance = []
referral_balance =[]
profit_and_loss_data = []
user_order_data = []
user_furture_count = []
user_manual_closing_count = []
user_forced_liquidation_count =[]
user_notClose_count =[]
########
ospath = os.path.abspath(os.path.dirname(os.getcwd()))
parser = argparse.ArgumentParser(description='參數專用')
parser.add_argument('--filename', type=str, default='20240605_2', help='文件名稱')
parser.add_argument('--env', type=str, default='test', help='用於要跑的環境')
args = parser.parse_args()
domain = 'http://' + args.env + '.top.one'
now = datetime.now()
dataetime = now.strftime('%H%M%S')
log_name_txt = ospath + '/logs/ReportLogName_tw.txt' 
output_file = ospath + '/reportExcel/' + str(datetime.now().strftime('%Y-%m-%d_%H%M%S')) + '_User_orders_Stats.xlsx'
#log_file = ospath + '/qa-simulated-trade-python/logs/20240605_2_171926_logs_tw.txt'
logReportFile = open(log_name_txt, 'r')
logName = logReportFile.readline()
log_path = ospath + '/logs/' + logName
# 从日志文件中提取UID的函数
def extract_uids(log_file):
    with open(log_file, 'r', encoding='utf-8') as file:
        logs = file.read()
    # with open(log_file, 'r', encoding='utf-8') as file:
    #     lines = file.readlines()
    # open_success_count = 0
    # close_success_count = 0
    # open_failure_count = 0
    # close_failure_count = 0
    # for line in lines:
    #     print(line)
    #     if "開倉成功" in line :
    #         open_success_count += 1
    #     elif "平倉成功" in line:
    #         close_success_count += 1
    #     elif "開倉失敗" in line:
    #         open_failure_count += 1
    #     elif "平倉失敗" in line:
    #         close_failure_count += 1
    #     else:
    #         continue
    # print('open_success_count:',open_success_count)
    # print('close_success_count:',close_success_count)
    # print('open_failure_count:',open_failure_count)
    # print('close_failure_count:',close_failure_count)
    # 使用正则表达式分割每个请求块
    log_entries = re.split(r'=+', logs)

    uids = set()  # 存储唯一的UID

    for entry in log_entries:
        # 提取UID
        uid_match = re.search(r'UID\(Top One\):(\w+)/', entry)
        uid = uid_match.group(1) if uid_match else None
        
        if uid:
            uids.add(uid)

    return list(uids)
#UID各種資料
def user_all_data(uids):
    for uid in uids:
        profit_and_loss = 0.0
        future_count = 0 #倉位數量
        manual_closing_count = 0 #手動平倉數量
        forced_liquidation_count = 0 #強制爆倉
        notClose_count = 0
        #close_count = 0
        user_data = {'倉位資訊':[]}
        balance_path = domain + '/wallet/v1/balance/' + uid   + '/user' #/v1/balance/:uid/user
        resp = rs.get(balance_path,headers=headers)
        if resp.status_code == 200:
            respJson = json.loads(resp.text)
            balance.append(respJson.get('data')[0].get('contract').get('available'))
        referral_balance_path =  domain + '/commission/v1/referral/wallet/balance/' + uid
        resp = rs.get(referral_balance_path,headers=headers)
        if resp.status_code == 200:
            respJson = json.loads(resp.text)
            referral_balance.append(respJson.get('data').get('data')[0].get('available'))
        future_orderList_path = domain + '/orchid/v1/future/position?uid=' + uid +'&page_size=1000'
        resp = rs.get(future_orderList_path,headers=headers)
        for list in json.loads(resp.text).get('data').get('list'):
            if list['close_type'] == 1:
                notClose_count += 1
            elif list['close_type'] == 2:
                manual_closing_count += 1
            elif list['close_type'] == 5:
                forced_liquidation_count += 1
            future_count += 1
            profit_and_loss = profit_and_loss + float(list['profit_and_loss'])
            # tmp = {'倉位ID': list['order_id'],'開倉價格': list['open_price']}
            # user_data['倉位資訊'].append(tmp)
        user_furture_count.append(future_count)
        profit_and_loss_data.append(profit_and_loss)
        # user_order_data.append(user_data)
        user_manual_closing_count.append(manual_closing_count)
        user_forced_liquidation_count.append(forced_liquidation_count)
        user_notClose_count.append(notClose_count)
# 创建Excel文件并填入提取的UID
def create_excel(uids,balances,referral_balances,profit_and_loss_datas,user_furture_counts,user_manual_closing_counts,user_forced_liquidation_counts,user_notClose_counts,output_file):
    wb = openpyxl.Workbook()
    # 处理UID数据
    ws_uid = wb.active
    ws_uid.title = "UID数据"

    headers = ["UID", "交易錢包餘額", "反擁錢包餘額", "已實現損益",'總開倉數量','持倉中數量','手動平倉數量','強制爆倉數量']
    # 添加标题
    for col_num, header in enumerate(headers, 1):
        ws_uid.cell(row=1, column=col_num, value=header)

    # 添加数据
    for row_num, uid in enumerate(uids, 2):
        ws_uid.cell(row=row_num, column=1, value=uid)
    for row_num, balance in enumerate(balances, 2):
        ws_uid.cell(row=row_num, column=2, value=balance)
    for row_num, referral_balance in enumerate(referral_balances, 2):
        ws_uid.cell(row=row_num, column=3, value=referral_balance)
    for row_num, profit_and_loss_data in enumerate(profit_and_loss_datas, 2):
        ws_uid.cell(row=row_num, column=4, value=profit_and_loss_data)
    for row_num, user_furture_count in enumerate(user_furture_counts, 2):
        ws_uid.cell(row=row_num, column=5, value=user_furture_count)
    for row_num, user_notClose_count in enumerate(user_notClose_counts, 2):
        ws_uid.cell(row=row_num, column=6, value=user_notClose_count)
    for row_num, user_manual_closing_count in enumerate(user_manual_closing_counts, 2):
        ws_uid.cell(row=row_num, column=7, value=user_manual_closing_count)
    for row_num, user_forced_liquidation_count in enumerate(user_forced_liquidation_counts, 2):
        ws_uid.cell(row=row_num, column=8, value=user_forced_liquidation_count)
    # 自动调整列宽
    for col in ws_uid.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws_uid.column_dimensions[column].width = adjusted_width

    # 保存文件
    wb.save(output_file)

# 主函数，用于处理日志文件并创建Excel文件
def main():
    #log_file = ospath +'/logs/2024-06-17_200752_logs_tw.txt'
    #output_file = '2024-06-17_uid_data.xlsx'
    uids = extract_uids(log_path)
    user_all_data(uids)
    create_excel(uids,balance,referral_balance,profit_and_loss_data,user_furture_count,user_manual_closing_count,user_forced_liquidation_count,user_notClose_count, output_file)
    print(f"Excel文件 '{output_file}' 创建成功。")
if __name__ == "__main__":
    main()
#/Users/hqcc-user32/Downloads/2024-06-14_195354_logs_tw.txt